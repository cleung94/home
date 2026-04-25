import re
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# -------------------------------------------------------------------
# Load the raw system exports and standards reference sheet.
# -------------------------------------------------------------------
ccure = pd.read_csv("ccure_export.csv")
genetec = pd.read_csv("genetec_export.csv")
standards = pd.read_csv("device_standards.csv")

# -------------------------------------------------------------------
# Set a fixed project date so the standards check is reproducible.
# -------------------------------------------------------------------
today = pd.Timestamp("2026-04-15")

# -------------------------------------------------------------------
# Build lookup dictionaries from the standards sheet.
# -------------------------------------------------------------------
standard_end_dates = standards.set_index("Model Name")["Out-of-Standard Date"].to_dict()
approved_firmware = standards.set_index("Model Name")["Approved Firmware Version"].to_dict()

# -------------------------------------------------------------------
# Helper function:
# Check whether the Programming Name follows the required format
# using only the Programming Name itself.
# -------------------------------------------------------------------
def is_valid_programming_name_format(programming_name):
    programming_name = str(programming_name).strip()
    pattern = r"^SITE\d{2} \([A-Z0-9\s\-/&]+\) [A-Z0-9\s\-/&]+ \[[A-Z]{2,4}\d{5}\]$"
    return bool(re.match(pattern, programming_name))

# -------------------------------------------------------------------
# Helper function:
# Check whether the Programming Name matches what it should be
# based on the row's Site Code, Location, Description, and
# Device Reference Number.
# -------------------------------------------------------------------
def does_programming_name_match_fields(site_code, location, description, reference_number, programming_name):
    programming_name = str(programming_name).strip()
    site_code = str(site_code).strip()
    location = str(location).strip()
    description = str(description).strip()
    reference_number = str(reference_number).strip()

    expected_name = f"{site_code} ({location}) {description} [{reference_number}]"
    return programming_name == expected_name

# -------------------------------------------------------------------
# Create validation flags for CCure on a working copy.
# -------------------------------------------------------------------
ccure_working = ccure.copy()

ccure_working["Issue_Online_Config_Flag"] = (
    (ccure_working["Online"] != "Online") |
    (ccure_working["Set of Configurations"] != "Configured Correctly")
).astype(int)

ccure_working["Issue_Name_Format_Flag"] = ccure_working["Programming Name"].apply(
    lambda name: 0 if is_valid_programming_name_format(name) else 1
)

ccure_working["Issue_Name_Match_Flag"] = ccure_working.apply(
    lambda row: 0 if does_programming_name_match_fields(
        row["Site Code"],
        row["Location"],
        row["Description"],
        row["Device Reference Number"],
        row["Programming Name"]
    ) else 1,
    axis=1
)

ccure_working["Issue_Standard_Flag"] = ccure_working["iSTAR Model"].apply(
    lambda model: 1 if pd.Timestamp(standard_end_dates[model]) <= today else 0
)

# -------------------------------------------------------------------
# Create validation flags for Genetec on a working copy.
# -------------------------------------------------------------------
genetec_working = genetec.copy()

genetec_working["Issue_Online_Config_Flag"] = (
    (genetec_working["Online"] != "Online") |
    (genetec_working["Set of Configurations"] != "Configured Correctly")
).astype(int)

genetec_working["Issue_Name_Format_Flag"] = genetec_working["Programming Name"].apply(
    lambda name: 0 if is_valid_programming_name_format(name) else 1
)

genetec_working["Issue_Name_Match_Flag"] = genetec_working.apply(
    lambda row: 0 if does_programming_name_match_fields(
        row["Site Code"],
        row["Location"],
        row["Description"],
        row["Device Reference Number"],
        row["Programming Name"]
    ) else 1,
    axis=1
)

genetec_working["Issue_Standard_Flag"] = genetec_working["Camera Model"].apply(
    lambda model: 1 if pd.Timestamp(standard_end_dates[model]) <= today else 0
)

genetec_working["Issue_IPv6_Flag"] = genetec_working["IP Address"].apply(
    lambda ip_address: 0 if ":" in str(ip_address) else 1
)

genetec_working["Issue_Offline_Flag"] = (genetec_working["Online"] != "Online").astype(int)

genetec_working["Issue_Firmware_Flag"] = genetec_working.apply(
    lambda row: 1 if row["Firmware Version"] != approved_firmware[row["Camera Model"]] else 0,
    axis=1
)

# -------------------------------------------------------------------
# Build the audit summary table.
# -------------------------------------------------------------------
answers = pd.DataFrame([
    {
        "Question": "Are all devices online and configured correctly?",
        "Status": "FAIL" if ccure_working["Issue_Online_Config_Flag"].sum() + genetec_working["Issue_Online_Config_Flag"].sum() > 0 else "PASS",
        "Issue Count": int(ccure_working["Issue_Online_Config_Flag"].sum() + genetec_working["Issue_Online_Config_Flag"].sum())
    },
    {
        "Question": "Do programming names follow the correct naming format?",
        "Status": "FAIL" if ccure_working["Issue_Name_Format_Flag"].sum() + genetec_working["Issue_Name_Format_Flag"].sum() > 0 else "PASS",
        "Issue Count": int(ccure_working["Issue_Name_Format_Flag"].sum() + genetec_working["Issue_Name_Format_Flag"].sum())
    },
    {
        "Question": "Do programming names match the expected device fields?",
        "Status": "FAIL" if ccure_working["Issue_Name_Match_Flag"].sum() + genetec_working["Issue_Name_Match_Flag"].sum() > 0 else "PASS",
        "Issue Count": int(ccure_working["Issue_Name_Match_Flag"].sum() + genetec_working["Issue_Name_Match_Flag"].sum())
    },
    {
        "Question": "Are all products currently used within our systems in-standard?",
        "Status": "FAIL" if ccure_working["Issue_Standard_Flag"].sum() + genetec_working["Issue_Standard_Flag"].sum() > 0 else "PASS",
        "Issue Count": int(ccure_working["Issue_Standard_Flag"].sum() + genetec_working["Issue_Standard_Flag"].sum())
    },
    {
        "Question": "Are all cameras in Genetec currently using IPv6?",
        "Status": "FAIL" if genetec_working["Issue_IPv6_Flag"].sum() > 0 else "PASS",
        "Issue Count": int(genetec_working["Issue_IPv6_Flag"].sum())
    },
    {
        "Question": "Are all cameras in Genetec online?",
        "Status": "FAIL" if genetec_working["Issue_Offline_Flag"].sum() > 0 else "PASS",
        "Issue Count": int(genetec_working["Issue_Offline_Flag"].sum())
    },
    {
        "Question": "Are all cameras in Genetec with the correct firmware version?",
        "Status": "FAIL" if genetec_working["Issue_Firmware_Flag"].sum() > 0 else "PASS",
        "Issue Count": int(genetec_working["Issue_Firmware_Flag"].sum())
    }
])

print("\nAudit answers\n")
print(answers.to_string(index=False))

# -------------------------------------------------------------------
# Helper function:
# Build a readable failure reason string for each failed device.
# -------------------------------------------------------------------
def build_failure_reason(row, system_name):
    reasons = []

    if row.get("Issue_Online_Config_Flag", 0) == 1:
        if row.get("Online") != "Online":
            reasons.append("Offline")
        if row.get("Set of Configurations") != "Configured Correctly":
            reasons.append("Configured Incorrectly")

    if row.get("Issue_Name_Format_Flag", 0) == 1:
        reasons.append("Programming Name Has Invalid Format")

    if row.get("Issue_Name_Match_Flag", 0) == 1:
        reasons.append("Programming Name Does Not Match Device Fields")

    if row.get("Issue_Standard_Flag", 0) == 1:
        reasons.append("Out of Standard")

    if system_name == "Genetec":
        if row.get("Issue_IPv6_Flag", 0) == 1:
            reasons.append("Using IPv4 Instead of IPv6")
        if row.get("Issue_Offline_Flag", 0) == 1 and "Offline" not in reasons:
            reasons.append("Offline")
        if row.get("Issue_Firmware_Flag", 0) == 1:
            reasons.append("Incorrect Firmware Version")

    return "; ".join(reasons)

# -------------------------------------------------------------------
# Build failed device tables WITHOUT modifying the raw exports.
# -------------------------------------------------------------------
ccure_failed = ccure_working.copy()
genetec_failed = genetec_working.copy()

ccure_failed["Reason for Failure"] = ccure_failed.apply(
    lambda row: build_failure_reason(row, "CCure"),
    axis=1
)

genetec_failed["Reason for Failure"] = genetec_failed.apply(
    lambda row: build_failure_reason(row, "Genetec"),
    axis=1
)

ccure_failed = ccure_failed[ccure_failed["Reason for Failure"] != ""].copy()
genetec_failed = genetec_failed[genetec_failed["Reason for Failure"] != ""].copy()

failure_columns = [
    "System",
    "Site Code",
    "Location",
    "Device Type",
    "Description",
    "Device Reference Number",
    "Programming Name",
    "Reason for Failure"
]

failed_devices = pd.concat(
    [ccure_failed[failure_columns], genetec_failed[failure_columns]],
    ignore_index=True
)

# -------------------------------------------------------------------
# Build the Audit Report sheet content.
# Audit Summary is included in this same tab.
# -------------------------------------------------------------------
author = "Christopher Leung"
report_date = today.strftime("%Y-%m-%d")

audit_scope = [
    "1. Device online and configuration status",
    "2. Programming name naming format compliance",
    "3. Programming name field match validation",
    "4. Product lifecycle compliance (in-standard vs out-of-standard)",
    "5. Genetec IPv6 usage",
    "6. Genetec camera online status",
    "7. Genetec firmware compliance"
]

audit_methodology = [
    "CCure and Genetec exports were loaded into pandas DataFrames.",
    "A standards reference sheet was used to validate approved models and firmware.",
    "Programming names were checked using regex pattern validation and field-level exact comparison.",
    "Online and configuration checks were performed using system export values.",
    "Genetec IP addresses were evaluated to determine whether they were IPv4 or IPv6.",
    "Issue flags were aggregated to produce an audit summary and a failed-devices remediation list."
]

key_results = [
    f"Total CCure device issues found: {int(ccure_working['Issue_Online_Config_Flag'].sum() + ccure_working['Issue_Name_Format_Flag'].sum() + ccure_working['Issue_Name_Match_Flag'].sum() + ccure_working['Issue_Standard_Flag'].sum())}",
    f"Total Genetec device issues found: {int(genetec_working['Issue_Online_Config_Flag'].sum() + genetec_working['Issue_Name_Format_Flag'].sum() + genetec_working['Issue_Name_Match_Flag'].sum() + genetec_working['Issue_Standard_Flag'].sum() + genetec_working['Issue_IPv6_Flag'].sum() + genetec_working['Issue_Offline_Flag'].sum() + genetec_working['Issue_Firmware_Flag'].sum())}",
    f"Total failed devices listed for remediation: {len(failed_devices)}"
]

cover_rows = [
    ["", ""],
    ["", ""],
    ["Author", author],
    ["Report Date", report_date],
    ["Report Type", "Python Validation Audit"],
    ["", ""],
    ["What Was Tested", ""],
]

for item in audit_scope:
    cover_rows.append(["", item])

cover_rows.append(["", ""])
cover_rows.append(["How It Was Tested", ""])

for item in audit_methodology:
    cover_rows.append(["", item])

cover_rows.append(["", ""])
cover_rows.append(["Key Results", ""])

for item in key_results:
    cover_rows.append(["", item])

cover_rows.append(["", ""])
cover_rows.append(["Audit Summary", ""])
cover_rows.append(["Question", "Status | Issue Count"])

for _, row in answers.iterrows():
    cover_rows.append([row["Question"], f'{row["Status"]} | {row["Issue Count"]}'])

cover_df = pd.DataFrame(cover_rows, columns=["Section", "Details"])

# -------------------------------------------------------------------
# Export the final workbook.
# -------------------------------------------------------------------
output_file = "access_control_audit_results.xlsx"

with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    ccure.to_excel(writer, sheet_name="CCure Export", index=False)
    genetec.to_excel(writer, sheet_name="Genetec Export", index=False)
    standards.to_excel(writer, sheet_name="Device Standards", index=False)
    cover_df.to_excel(writer, sheet_name="Audit Report", index=False)
    failed_devices.to_excel(writer, sheet_name="Failed Devices", index=False)

# -------------------------------------------------------------------
# Formatting helpers
# -------------------------------------------------------------------
def autosize_columns(ws):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_length = 0
        for cell in col_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 60)

def estimate_row_height(ws, row_idx, max_cols_to_check=None):
    base_height = 18
    per_line_height = 15
    max_lines = 1

    cells = list(ws[row_idx])
    if max_cols_to_check is not None:
        cells = cells[:max_cols_to_check]

    for cell in cells:
        value = "" if cell.value is None else str(cell.value)
        col_letter = get_column_letter(cell.column)
        width = ws.column_dimensions[col_letter].width or 10

        if width <= 0:
            width = 10

        manual_lines = value.count("\n") + 1
        estimated_wrap_lines = max(1, math.ceil(len(value) / max(width - 2, 1)))
        lines = max(manual_lines, estimated_wrap_lines)
        max_lines = max(max_lines, lines)

    ws.row_dimensions[row_idx].height = max(base_height, min(base_height + (max_lines - 1) * per_line_height, 90))

def hide_blank_columns(ws):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        has_content = any(cell.value not in (None, "") for cell in col_cells)
        if not has_content:
            ws.column_dimensions[col_letter].hidden = True

# -------------------------------------------------------------------
# Apply formatting with openpyxl.
# -------------------------------------------------------------------
workbook = load_workbook(output_file)

dark_blue_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
medium_blue_fill = PatternFill(fill_type="solid", fgColor="2F75B5")
section_fill = PatternFill(fill_type="solid", fgColor="B4C7E7")

white_bold_font = Font(bold=True, color="FFFFFF", size=14)
header_font = Font(bold=True)
white_header_font = Font(bold=True, color="FFFFFF")
normal_font = Font(bold=False)

center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

for sheet_name in ["CCure Export", "Genetec Export", "Device Standards", "Failed Devices"]:
    ws = workbook[sheet_name]

    for cell in ws[1]:
        cell.font = white_header_font
        cell.fill = dark_blue_fill
        cell.alignment = center_wrap

    for row in range(2, ws.max_row + 1):
        for cell in ws[row]:
            cell.alignment = left_wrap

    autosize_columns(ws)
    hide_blank_columns(ws)

    for row in range(1, ws.max_row + 1):
        estimate_row_height(ws, row)

    ws.freeze_panes = "A2"

report_ws = workbook["Audit Report"]

report_ws.merge_cells("A1:B1")
report_ws["A1"] = "Physical Security Systems and Health Configuration Report"
report_ws["A1"].font = white_bold_font
report_ws["A1"].fill = dark_blue_fill
report_ws["A1"].alignment = center_wrap

for cell in report_ws[2]:
    cell.value = None
    cell.fill = PatternFill(fill_type=None)
    cell.font = normal_font
    cell.alignment = left_wrap

for cell in report_ws[3]:
    cell.value = None
    cell.fill = PatternFill(fill_type=None)
    cell.font = normal_font
    cell.alignment = left_wrap

for row in range(4, report_ws.max_row + 1):
    a_val = report_ws[f"A{row}"].value
    b_val = report_ws[f"B{row}"].value

    report_ws[f"A{row}"].alignment = left_wrap
    report_ws[f"B{row}"].alignment = left_wrap

    if a_val in ["Author", "Report Date", "Report Type"]:
        report_ws[f"A{row}"].font = header_font

    if a_val in ["What Was Tested", "How It Was Tested", "Key Results", "Audit Summary"]:
        report_ws[f"A{row}"].font = header_font
        report_ws[f"A{row}"].fill = section_fill
        report_ws[f"B{row}"].fill = section_fill

    if a_val == "Question" and b_val == "Status | Issue Count":
        report_ws[f"A{row}"].font = white_header_font
        report_ws[f"B{row}"].font = white_header_font
        report_ws[f"A{row}"].fill = medium_blue_fill
        report_ws[f"B{row}"].fill = medium_blue_fill
        report_ws[f"A{row}"].alignment = center_wrap
        report_ws[f"B{row}"].alignment = center_wrap

for row in range(1, report_ws.max_row + 1):
    for cell in report_ws[row]:
        if row == 1:
            cell.alignment = center_wrap
        else:
            cell.alignment = left_wrap

autosize_columns(report_ws)
hide_blank_columns(report_ws)

report_ws.row_dimensions[1].height = 28

for row in range(1, report_ws.max_row + 1):
    estimate_row_height(report_ws, row, max_cols_to_check=2)

report_ws.freeze_panes = "A4"

workbook.save(output_file)

print(f"\nSaved {output_file}")
print(f"Failed device count: {len(failed_devices)}")
